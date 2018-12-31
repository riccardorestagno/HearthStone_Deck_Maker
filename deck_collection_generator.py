from bs4 import BeautifulSoup
import requests
import openpyxl
from openpyxl.styles import Font
from tkinter import *
import os
import platform


def soup_session(link):
    """BeautifulSoup session"""
    session = requests.Session().get(link)
    soup = BeautifulSoup(session.content, 'html.parser')
    return soup


def get_card_info(card):

    # Converts card type NONE to Neutral if present
    card_class = card['data-card-class']
    if card_class == "NONE":
        card_class = "Neutral"

    # Gets card mana cost
    card_mana_cost = card['data-card-mana-cost']

    # Gets cards name
    for card_art in card.find_all('img'):
        card_name = '=HYPERLINK("{}", "{}")'.format(card_art['data-src'], card['data-card-name'])

    # Gets quantity of card owned
    for card_count in card.find_all('span'):
        try:
            amount_owned = card_count['data-card-count']
            break
        except KeyError:
            continue

    # Gets cards rarity
    rarity_dict = {"1": "Common", "2": "None", "3": "Rare", "4": "Epic", "5": "Legendary"}
    card_rarity = rarity_dict[card['data-rarity']]

    # Shows if card is golden or not
    if card['data-is-gold'] == 'True':
        is_card_golden = 'True'
    else:
        is_card_golden = ''

    return card_class, card_mana_cost, card_name, amount_owned, card_rarity, is_card_golden


def get_collection(session):

    my_card_collection = []

    for card in session.find_all('div', attrs={'class': 'card-image-item owns-card'}):
        card_class, card_mana_cost, card_name, amount_owned, card_rarity, is_card_golden = get_card_info(card)
        my_card_collection.append([card_class, card_mana_cost, card_name, amount_owned, card_rarity, is_card_golden])

    return my_card_collection


def format_cells(workbook):

    length = 0

    def as_text(value):
        if value is None:
            return ""
        return str(value)

    for sheet in workbook.worksheets:
        for column_cells in sheet.columns:
            for cell in column_cells:
                if "HYPERLINK" not in as_text(cell.value):
                    new_length = len(as_text(cell.value))
                else:
                    new_length = len(cell.value.split('", "')[1][:-2])
                length = max(new_length, length)
            sheet.column_dimensions[column_cells[0].column].width = max(10, length) * 1.1
            length = 0

    return workbook


def create_excel_file(collection, file_path):

    card_classes = ["Druid", "Hunter", "Mage", "Paladin", "Priest", "Rogue", "Shaman", "Warlock", "Warrior", "Neutral"]
    workbook = openpyxl.Workbook()

    # remove initial sheet
    first_sheet = workbook["Sheet"]
    workbook.remove(first_sheet)

    # Creates organized tabs with each card
    for card_class in card_classes:
        workbook.create_sheet().title = card_class
        font = Font(bold=True)
        workbook[card_class].append(["Mana Cost", "Card Name", "Card Quantity", "Rarity", "Is Golden"])
        for cell in workbook[card_class]["1:1"]:
            cell.font = font

    for card in collection:
        workbook[card[0].title()].append([card[1], card[2], card[3], card[4], card[5]])

    hyperlink = Font(underline='single', color='0563C1')
    for card_class in card_classes:
        worksheet = workbook[card_class]
        for row in worksheet['B2:B{}'.format(worksheet.max_row)]:
            for cell in row:
                cell.font = hyperlink

    workbook = format_cells(workbook)

    workbook.save(file_path)


def run_spreadsheet_creator(hearthpwn_card_collection_url, file_path):

    profile_session = soup_session(hearthpwn_card_collection_url)
    card_collection = get_collection(profile_session)
    create_excel_file(card_collection, file_path)


def tkinter_input_validation(hearthpwn_username, hearthpwn_card_collection_url, file_path):

    throw_error = False
    invalid_chars = ['\\', '/', ':', '?', '|', '<', '>', '"', '*']

    if not hearthpwn_username:
        activate_error_dialog_box("The username field is required.")
        return True

    if not file_path or file_path.lower().replace('.xlsx', '').endswith('/') \
            or file_path.lower().replace('.xlsx', '').endswith('\\'):
        activate_error_dialog_box("The file path field with a valid file name is required.")
        return True

    if (platform.system() == "Windows" and any(substring in file_path.rsplit('\\')[-1] for substring in invalid_chars)
            or platform.system() != "Windows" and any(substring in file_path.rsplit('/')[-1] for substring in invalid_chars)):
        activate_error_dialog_box(
            """A file name cannot contain any of the following characters: '\\', '/', ':', '?', '|', '<', '>', '"', '*'""")
        return True

    if (platform.system() == "Windows" and "\\" in file_path and not os.path.exists(file_path.rsplit("\\")[0]))\
            or ("/" in file_path and not os.path.exists(file_path.rsplit("/")[0])):
        activate_error_dialog_box("The file path entered does not exist.")
        return True

    session = soup_session(hearthpwn_card_collection_url)
    if 'not found - hearthpwn' in session.find('title').text.lower():
        activate_error_dialog_box("An invalid or non-public HearthPwn username was entered.")
        return True

    return throw_error


def activate_filepath_dialog_box(file_path):

    root = Tk()
    root.title("Excel File Created!")
    root.geometry("450x150")

    file_created_text = StringVar()
    file_created_text.set("HearthStone excel file successfully created! It can be found at: " + file_path)
    file_created_box = Label(root, textvariable=file_created_text)
    file_created_box.config(font=('Calibri', 16), wraplength=400, justify=CENTER)
    file_created_box.pack()

    close = Button(root, text="Close", command=root.destroy)
    close.pack(side=BOTTOM)
    root.mainloop()


def activate_error_dialog_box(error_message):

    def return_to_main_screen():
        root.destroy()
        activate_start_widget()

    root = Tk()
    root.title("An Error Occured")
    root.geometry("450x150")

    error_text = StringVar()
    error_text.set(error_message)
    error_box = Label(root, textvariable=error_text)
    error_box.config(font=('Calibri', 16), wraplength=400, justify=CENTER)
    error_box.pack()

    try_again = Button(root, text="Try Again", command=return_to_main_screen)
    try_again.pack(side=BOTTOM)
    root.mainloop()


def activate_start_widget():

    def on_button_click():
        file_path = file_path_to_save_entry.get()
        hearthpwn_username = username_entry.get()
        root.destroy()
        hearthpwn_card_collection_url = "https://www.hearthpwn.com/members/{}/collection".format(hearthpwn_username)

        # VALIDATION
        throw_error = tkinter_input_validation(hearthpwn_username, hearthpwn_card_collection_url, file_path)

        if not throw_error:

            if not file_path.lower().endswith('.xlsx'):
                file_path += '.xlsx'
            try:
                run_spreadsheet_creator(hearthpwn_card_collection_url, file_path)
                if any(substring in file_path for substring in ['/', '\\']):
                    activate_filepath_dialog_box(file_path)
                else:
                    if platform.system() == "Windows":
                        file_path_char = '\\'
                    else:
                        file_path_char = '/'
                    activate_filepath_dialog_box(os.getcwd() + file_path_char + file_path)
            except PermissionError:
                activate_error_dialog_box("Unable to create new file with the name provided "
                                          "since a file with that name is currently in use.")

    root = Tk()
    root.title("Create Excel File of HearthStone Collection")
    root.geometry("450x150")
    username = Label(root, text="HearthPwn Username")
    username_entry = Entry(root, bd=5)

    file_path_to_save = Label(root, text="Desired File Path (Including File Name)")
    file_path_to_save_entry = Entry(root, bd=5, width=50)

    submit = Button(root, text="Submit", command=on_button_click)

    username.pack()
    username_entry.pack()
    file_path_to_save.pack()
    file_path_to_save_entry.pack()
    submit.pack(side=BOTTOM)
    root.mainloop()


if __name__ == "__main__":
    activate_start_widget()
